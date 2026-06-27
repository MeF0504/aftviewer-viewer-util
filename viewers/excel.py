from __future__ import annotations

import argparse
from pathlib import Path
from logging import getLogger
from functools import partial

from aftviewer import (GLOBAL_CONF, Args, print_error, print_warning,
                       print_key, interactive_cui, help_template,
                       add_args_encoding, add_args_specification,
                       ReturnMessage as RM)

if 'tabulate' in GLOBAL_CONF.pack_list:
    from tabulate import tabulate
    is_tabulate = True
else:
    print("I can't find tabulate library.")
    is_tabulate = False
logger = getLogger(GLOBAL_CONF.logname)
logger.info(f'use tabulate: {is_tabulate}')


def add_args(parser: argparse.ArgumentParser) -> None:
    parser.add_argument('--ext', '-e', choices=['xls', 'xlsx', 'xlsm'],
                        help='Specify the version of the Excel file.',
                        default=None)
    add_args_encoding(parser, help='Override the encoding of the ".xls" file.')
    kwargs_k = dict(help='Specify the sheet names to show.'
                    ' If not specified, all sheets will be shown.'
                    ' If empty, all sheet names will be printed.')
    add_args_specification(parser, verbose=False, key=True,
                           interactive=False, cui=True,
                           kwargs_k=kwargs_k)


def show_help() -> None:
    helpmsg = help_template('excel', 'Show the contents of Excel files.',
                            add_args)
    print(helpmsg)


def get_contents(allsheets: list[str], path: Path) -> tuple[list, list]:
    is_root = str(path) == '.'
    logger.debug(f'path: {path} : {is_root}')
    if is_root:
        # at root
        return [], sorted(allsheets)
    else:
        return [], []


def get_sheets(allsheets: list[str], args: Args) -> None | list[str]:
    if args.key is None:
        return allsheets
    elif len(args.key) == 0:
        [print(s) for s in allsheets]
        return None

    res = []
    for k in args.key:
        if k not in allsheets:
            print_warning(f'Sheet "{k}" not found in the Excel file.')
        else:
            res.append(k)
    return res


def show_table(data: dict[str, list[list[str]]], table_path: str,
               **kwargs) -> RM:
    values = data[table_path]
    if len(values) == 0:
        logger.debug(f'empty table ({table_path}): {values}')
        return RM('No data in the table.', False)
    alp = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    headers = ['']
    logger.debug(f'cpath: {table_path}, cols: {len(values[0])}')
    assert len(values[0]) < len(alp)**2, \
        f'Too many columns in the values: {len(values[0])}. '
    for i in range(len(values[0])):
        if i < len(alp):
            headers.append(alp[i])
        else:
            headers.append(f'{alp[i // len(alp)]}{alp[i % len(alp)]}')
    table_data = [[f'{i+1}'] + row for i, row in enumerate(values)]
    if is_tabulate:
        table_str = tabulate(table_data, headers, tablefmt='simple_grid')
    else:
        table_str = ' | '.join(headers) + '\n'
        for row in table_data:
            row_str = ' | '.join(str(cell) for cell in row)
            table_str += row_str + '\n'
    return RM(table_str, False)


def get_data_xls(fpath: Path, args: Args) -> None | dict[str, list[list[str]]]:
    import xlrd
    book = xlrd.open_workbook(fpath, encoding_override=args.encoding)
    sheets = get_sheets(book.sheet_names(), args)
    if sheets is None:
        # showing all sheets case.
        return None

    res = {}
    for sh in sheets:
        ws = book.sheet_by_name(sh)
        vals = []
        for row_idx in range(ws.nrows):
            row = ws.row_values(row_idx)
            row2 = [str(cell) if cell is not None else '' for cell in row]
            vals.append(row2)
        res[sh] = vals
    return res


def get_data_xlsx(fpath: Path, args: Args,
                  keep_vba: bool) -> None | dict[str, list[list[str]]]:
    import openpyxl
    book = openpyxl.load_workbook(fpath, read_only=True, keep_vba=keep_vba)
    sheets = get_sheets(book.sheetnames, args)
    if sheets is None:
        # showing all sheets case.
        return None

    res = {}
    for sh in sheets:
        ws = book[sh]
        vals: list[list[str]] = []
        res[sh] = vals
        if not hasattr(ws, 'iter_rows'):
            logger.warning(f'Sheet "{sh}" does not support iter_rows')
            continue
        for row in ws.iter_rows(values_only=True):
            row2 = [str(cell) if cell is not None else '' for cell in row]
            vals.append(row2)
    return res


def main(fpath: Path, args: Args) -> int:
    if args.ext is None:
        excelver = fpath.suffix.lower()[1:]
    else:
        excelver = args.ext

    if excelver not in ['xls', 'xlsx', 'xlsm']:
        print_error(f'Unsupported Excel file version: {excelver}')
        return 2
    if excelver == 'xls':
        data = get_data_xls(fpath, args)
    elif excelver == 'xlsx':
        data = get_data_xlsx(fpath, args, keep_vba=False)
    elif excelver == 'xlsm':
        data = get_data_xlsx(fpath, args, keep_vba=True)
    else:
        logger.error(f'Unsupported Excel version?? {excelver}')
        return 2
    if data is None:
        # showing all sheets case.
        return 0

    if args.cui:
        gc = partial(get_contents, list(data.keys()))
        sf = partial(show_table, data)
        interactive_cui(fpath.name, gc, sf)
        pass
    else:
        for sheet in data:
            print_key(sheet)
            print(show_table(data, sheet).message)

    return 0
